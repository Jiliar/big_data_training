import os

#XLS y XLSX
# Lectura de un archivo XLSX y creación de un diccionario con las subvenciones por asociación
from openpyxl import load_workbook, Workbook

base_dir = os.path.dirname(__file__)
ruta_entrada = os.path.join(base_dir, '..', '..', 'data', 'chapter1', 'subvenciones.xlsx')
ruta_salida = os.path.join(base_dir, '..', '..', 'data', 'chapter1', 'subvenciones_actualizadas.xlsx')

# Leer archivo .xlsx
wb = load_workbook(ruta_entrada)
ws = wb.active

asocs = {}

# Saltar encabezado y leer datos
for row in ws.iter_rows(min_row=2, values_only=True):
    centro = row[0]  # Asociación
    importe = row[2]  # Importe
    
    if centro is None or importe is None:
        continue  # Saltar filas con datos vacíos

    try:
        importe = float(str(importe).replace(',', '.'))  # Maneja números con comas
    except ValueError:
        print(f"Valor no válido en la fila: {row}")
        continue  # Ignorar filas con errores de conversión

    asocs[centro] = asocs.get(centro, 0) + importe

# Escribir archivo nuevo
wb_salida = Workbook()
ws_salida = wb_salida.active
ws_salida.title = 'Resumen Subvenciones'

ws_salida.append(['Asociación', 'Total Subvención'])

for centro, total in asocs.items():
    ws_salida.append([centro, total])

wb_salida.save(ruta_salida)
print("Archivo .xlsx generado:", ruta_salida)


# Lectura de un archivo XLS y creación de un diccionario con las subvenciones por asociación
import xlrd
import xlwt

ruta_entrada = os.path.join(base_dir, '..', '..', 'data', 'chapter1', 'subvenciones.xls')
ruta_salida = os.path.join(base_dir, '..', '..', 'data', 'chapter1', 'subvenciones_actualizadas.xls')

# Leer archivo .xls
libro = xlrd.open_workbook(ruta_entrada)
hoja = libro.sheet_by_index(0)

asocs = {}

# Saltar encabezado y leer datos
for i in range(1, hoja.nrows):
    centro = hoja.cell_value(i, 0)  # Suponiendo que columna 0 es 'Asociación'
    importe = float(hoja.cell_value(i, 2))  # Columna 2 es 'Importe'
    asocs[centro] = asocs.get(centro, 0) + importe

# Crear nuevo archivo con los resultados
libro_salida = xlwt.Workbook()
hoja_salida = libro_salida.add_sheet('Resumen Subvenciones')

hoja_salida.write(0, 0, 'Asociación')
hoja_salida.write(0, 1, 'Total Subvención')

for i, (centro, total) in enumerate(asocs.items(), start=1):
    hoja_salida.write(i, 0, centro)
    hoja_salida.write(i, 1, total)

libro_salida.save(ruta_salida)
print("Archivo .xls generado:", ruta_salida)

# Operación con Pandas:
import os
import pandas as pd

# Definir rutas
base_dir = os.path.dirname(__file__)
ruta_entrada = os.path.join(base_dir, '..', '..', 'data', 'chapter1', 'subvenciones.xlsx')
ruta_salida = os.path.join(base_dir, '..', '..', 'data', 'chapter1', 'subvenciones_actualizadas.xlsx')

# Leer archivo Excel
df = pd.read_excel(ruta_entrada)

# Limpiar y convertir la columna 'Importe' si tiene comas
df['Importe'] = pd.to_numeric(df['Importe'].astype(str).str.replace(',', '.'), errors='coerce')

# Eliminar filas con valores faltantes
df = df.dropna(subset=['Asociación', 'Importe'])

# Agrupar por asociación y sumar
df_total = df.groupby('Asociación', as_index=False)['Importe'].sum()
df_total.rename(columns={'Importe': 'Total Subvención'}, inplace=True)

# Guardar a un nuevo archivo Excel
df_total.to_excel(ruta_salida, index=False)
print("Archivo .xlsx generado:", ruta_salida)